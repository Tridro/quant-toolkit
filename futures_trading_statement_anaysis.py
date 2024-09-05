#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time         : 2019/10/16 11:04
# @Author       : Tridro
# @E-mail       : tridro@beneorigin.com
# @File         : futures_trading_statement_anaysis.py
# All CopyRight Reserved

import getopt
import os.path
import re
import sys
from datetime import datetime
from itertools import islice
from typing import Literal, Optional, Union

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from openpyxl.chart import Reference, LineChart, PieChart, BarChart, RadarChart
from openpyxl.chart.axis import NumericAxis, TextAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout

# ---------------------------------------------------- 基础数据 开始 ----------------------------------------------------
BASE_DIR = os.path.dirname(os.path.realpath(sys.argv[0]))

RISK_FREE_INTEREST_RATE: float = 0.0345
STATISTIC_METHOD: str = 'log'

CONTRACT_CODE = {'if': '沪深300股指', 'ih': '上证50股指', 'ic': '中证500股指', 'im': '中证1000股指', 'tf': '五债', 't': '十债',
                 'ts': '二债', 'tl': '三十债', 'cu': '铜', 'al': '铝', 'zn': '锌', 'pb': '铅', 'ni': '镍', 'sn': '锡',
                 'au': '黄金', 'ag': '白银', 'j': '焦炭', 'jm': '焦煤', 'zc': '动力煤ZC', 'rb': '螺纹钢', 'i': '铁矿石',
                 'hc': '热轧卷板', 'sf': '硅铁', 'sm': '锰硅', 'fg': '玻璃', 'ss': '不锈钢', 'wr': '线材', 'ru': '天然橡胶',
                 'sp': '漂针浆', 'bb': '细木工板', 'fb': '纤维板', 'nr': '20号胶', 'fu': '燃料油', 'bu': '石油沥青',
                 'l': '线型低密度聚乙烯', 'pp': '聚丙烯', 'v': '聚氯乙烯', 'ta': '精对苯二甲酸', 'ma': '甲醇MA', 'eg': '乙二醇',
                 'eb': '苯乙烯', 'ur': '尿素', 'sa': '纯碱', 'pg': '液化石油气', 'lu': '低硫燃料油', 'm': '豆粕', 'y': '豆油',
                 'oi': '菜籽油', 'a': '黄大豆1号', 'b': '黄大豆2号', 'p': '棕榈油', 'c': '黄玉米', 'rm': '菜籽粕',
                 'cs': '玉米淀粉', 'cf': '一号棉花', 'cy': '棉纱', 'sr': '白砂糖', 'wh': '强筋小麦', 'ri': '旱籼稻',
                 'rr': '粳米', 'rs': '油菜籽', 'jr': '粳稻谷', 'lr': '晚籼稻', 'pm': '普通小麦', 'sc': '原油', 'ap': '鲜苹果',
                 'jd': '鲜鸡蛋', 'cj': '干制红枣', 'pf': '短纤', 'bc': '国际铜', 'lh': '生猪', 'pk': '花生', 'si': '工业硅',
                 'ao': '氧化铝', 'lc': '碳酸锂', 'br': 'BR橡胶', 'ec': '集运欧线', 'px': '对二甲苯', 'sh': '烧碱', 
                 'pr': '瓶片'}

TRADING_UNIT = {'if': 300, 'ih': 300, 'ic': 200, 'im': 200, 'tf': 10000, 't': 10000, 'ts': 20000, 'tl': 10000, 'cu': 5,
                'al': 5, 'zn': 5, 'sc': 1000, 'pb': 5, 'ni': 1, 'sn': 1, 'au': 1000, 'ag': 15, 'j': 100, 'jm': 60,
                'zc': 100, 'rb': 10, 'i': 100, 'hc': 10, 'sf': 5, 'sm': 5, 'wr': 10, 'fu': 10, 'bu': 10, 'ru': 10,
                'l': 5, 'pp': 5, 'v': 5, 'ta': 5, 'ma': 10, 'sp': 10, 'm': 10, 'y': 10, 'oi': 10, 'a': 10, 'b': 10,
                'p': 10, 'c': 10, 'rm': 10, 'cs': 10, 'jd': 10, 'bb': 500, 'fb': 500, 'cf': 5, 'cy': 5, 'sr': 10,
                'wh': 20, 'ri': 20, 'jr': 20, 'lr': 20, 'fg': 20, 'ss': 5, 'nr': 10, 'eg': 10, 'eb': 5, 'ur': 20,
                'rr': 10, 'rs': 10, 'ap': 10, 'cj': 5, 'pm': 50, 'sa': 20, 'pg': 20, 'lu': 10, 'pf': 5, 'bc': 5,
                'lh': 16, 'pk': 5, 'si': 5, 'ao': 20, 'lc': 1, 'br': 5, 'ec': 50, 'px': 5, 'sh': 30, 'pr': 15}


# ---------------------------------------------------- 基础数据 结束 ----------------------------------------------------


# ---------------------------------------------------- 数据读取 开始 ----------------------------------------------------
def read_statement_files(dir_input=''):
    print(f"+{'-' * 23}  使用准备  {'-' * 23}+\n" +
          f"|{' ':58}|\n" +
          f"|{'请把所有交易结算单txt文件放入当前或指定路径的文件夹中!':^33}|\n" +
          f"|{' ':58}|\n" +
          f"+{'-' * 58}+\n")
    if os.path.isdir(dir_input):
        if os.path.exists(dir_input):
            path = dir_input
            folder_name = os.path.basename(path)
        else:
            input(f"\n{datetime.now()} | 错误 | {dir_input} 路径错误，请检查路径并重试，按任意键退出！\n")
            raise SystemExit()
    else:
        folder = input("输入文件夹名/文件夹路径：") if dir_input == '' else dir_input
        path = os.path.join(BASE_DIR, folder) if not os.path.isdir(folder) else folder
        folder_name = os.path.basename(path)
        if not os.path.exists(path):
            input(f"\n{datetime.now()} | 错误 | 找不到 {folder_name} 文件夹，请检查并重试，按任意键退出！\n")
            raise SystemExit()
    files_list = os.listdir(path)
    statement_data_list = []
    for file in files_list:
        file_dir = os.path.join(path, file)
        if os.path.isfile(file_dir):
            try:
                with open(file_dir) as f:
                    statement_data_list.append(f.readlines())
            except UnicodeDecodeError:
                with open(file_dir, encoding='gb18030') as f:
                    statement_data_list.append(f.readlines())
    print(f'\n{datetime.now()} | 信息 | 已读取 {folder_name} 文件夹')
    return statement_data_list


# ---------------------------------------------------- 数据读取 结束 ----------------------------------------------------


# ---------------------------------------------------- 数据提取 开始 ----------------------------------------------------
def data_extract(source, client_id=''):
    def regular_expression_search(file, expression, start=0):
        for ri in range(start, len(file)):
            matched_obj = re.search(fr'{expression}', file[ri])
            if matched_obj is not None:
                return matched_obj

    if client_id == '':
        client_id = regular_expression_search(source[0], r'^(客户号 Client ID)[：\s]*([0-9]+)').groups()[1]
    print(f'{datetime.now()} | 信息 | 开始客户号 {client_id} 的结算数据提取')
    account = pd.DataFrame(columns=['日期', '期初结存', '出入金', '平仓盈亏', '持仓盯市盈亏', '手续费', '期末结存',
                                    '客户权益', '保证金占用', '可用资金', '风险度', '权利金收入', '权利金支出', '多头期权市值',
                                    '空头期权市值', '市值权益'])
    transaction_record = pd.DataFrame(columns=['成交日期', '交易所', '品种', '合约', '买/卖', '投/保', '成交价', '手数',
                                               '成交额', '开/平', '手续费', '平仓盈亏', '权利金收支', '成交序号'])
    position_closed = pd.DataFrame(columns=['平仓日期', '交易所', '品种', '合约', '开仓日期', '买/卖', '手数', '开仓价',
                                            '昨结算', '成交价', '平仓盈亏', '权利金收支', '交易盈亏', '盈亏率'])
    sep = re.compile(r'[|\s|]+')

    for i in range(len(source)):
        statement_date = regular_expression_search(source[i], r'^(日期 Date)[：\s]*([0-9]+)').groups()[1]
        is_new_version = pd.to_datetime(statement_date) >= datetime(2022, 9, 23) or False
        file_client_id = regular_expression_search(source[i], r'^(客户号 Client ID)[：\s]*([0-9]+)').groups()[1]
        if file_client_id != client_id:
            input(f"{datetime.now()} | 错误 | {statement_date} 结算单客户号不正确, 请检查后重试, 按任意键退出!\n")
            raise SystemExit()
        print(f'\r{datetime.now()} | 信息 | 提取 {statement_date} 结算数据', end='')
        for j in range(len(source[i])):
            if re.match(r'\s*资金状况', source[i][j]):
                client_equity = float(regular_expression_search(source[i],
                                                                r'(客户权益 Client Equity)[：\s]*([0-9\.]+)',
                                                                start=j).groups()[1])
                margin_occupied = float(regular_expression_search(source[i],
                                                                  r'(保证金占用 Margin Occupied)[：\s]*([0-9\.]+)',
                                                                  start=j).groups()[1])
                data = {'日期': pd.to_datetime(statement_date),
                        '期初结存': float(
                            regular_expression_search(source[i], r'(期初结存 Balance b/f)[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '出入金': float(
                            regular_expression_search(source[i], r'(出 入 金 Deposit/Withdrawal)[：\s]*([-0-9\.]+)',
                                                      start=j).groups()[1]),
                        '平仓盈亏': float(
                            regular_expression_search(source[i], r'(平仓盈亏 Realized P/L)[：\s]*([-0-9\.]+)',
                                                      start=j).groups()[1]),
                        '持仓盯市盈亏': float(
                            regular_expression_search(source[i], r'(持仓盯市盈亏 MTM P/L)[：\s]*([-0-9\.]+)',
                                                      start=j).groups()[1]),
                        '手续费': float(
                            regular_expression_search(source[i], r'(手 续 费 Commission)[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '期末结存': float(
                            regular_expression_search(source[i], r'(期末结存 Balance c/f)[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '客户权益': client_equity,
                        '保证金占用': margin_occupied,
                        '可用资金': float(
                            regular_expression_search(source[i], r'(可用资金 Fund Avail\.)[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '风险度': margin_occupied / client_equity,
                        '权利金收入': float(
                            regular_expression_search(source[i], r'(权利金收入 Premium received)[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '权利金支出': float(
                            regular_expression_search(source[i], r'(权利金支出 Premium paid)[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '多头期权市值': float(
                            regular_expression_search(source[i], r'(多头期权市值 Market value\(long\))[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '空头期权市值': float(
                            regular_expression_search(source[i], r'(空头期权市值 Market value\(short\))[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1]),
                        '市值权益': float(
                            regular_expression_search(source[i], r'(市值权益 Market value\(equity\))[：\s]*([0-9\.]+)',
                                                      start=j).groups()[1])}
                df = pd.DataFrame([data])
                account = pd.concat([account, df], ignore_index=True)
            if re.match(r'\s*成交记录 Transaction Record', source[i][j]):
                separator_count = 0
                for ldx in range(j, len(source[i])):
                    if re.match(r'^\n', source[i][ldx]):
                        continue
                    if re.match(r'^-', source[i][ldx]):
                        separator_count += 1
                        if separator_count == 2:
                            continue
                        elif separator_count == 3:
                            break
                    if separator_count == 2:
                        row = sep.split(source[i][ldx])
                        data = {'成交日期': pd.to_datetime(row[1]),
                                '交易所': row[3] if is_new_version else row[2],
                                '品种': row[5] if is_new_version else row[3],
                                '合约': row[6] if is_new_version else row[4],
                                '买/卖': row[7] if is_new_version else row[5],
                                '投/保': row[8] if is_new_version else row[6],
                                '成交价': float(row[9]) if is_new_version else float(row[7]),
                                '手数': int(row[10]) if is_new_version else int(row[8]),
                                '成交额': float(row[11]) if is_new_version else float(row[9]),
                                '开/平': row[12] if is_new_version else row[10],
                                '手续费': float(row[13]) if is_new_version else float(row[11]),
                                '平仓盈亏': float(row[14]) if is_new_version else float(row[12]),
                                '权利金收支': float(row[15]) if is_new_version else float(row[13]),
                                '成交序号': int(row[16]) if is_new_version else int(row[14])}
                        df = pd.DataFrame([data])
                        transaction_record = pd.concat([transaction_record, df], ignore_index=True)
            if re.match(r'\s*平仓明细 Position Closed', source[i][j]):
                separator_count = 0
                for ldx in range(j, len(source[i])):
                    if re.match(r'^\n', source[i][ldx]):
                        continue
                    if re.match(r'^-', source[i][ldx]):
                        separator_count += 1
                        if separator_count == 2:
                            continue
                        elif separator_count == 3:
                            break
                    if separator_count == 2:
                        row = sep.split(source[i][ldx])
                        instrument = row[6] if is_new_version else row[4]
                        b_s = row[9] if is_new_version else row[6]
                        lots = int(row[10]) if is_new_version else int(row[7])
                        open_price = float(row[11]) if is_new_version else float(row[8])
                        transaction_price = float(row[13]) if is_new_version else float(row[10])
                        price_margin = transaction_price - open_price if b_s == '卖' else open_price - transaction_price
                        data = {'平仓日期': pd.to_datetime(row[1]),
                                '交易所': row[3] if is_new_version else row[2],
                                '品种': row[5] if is_new_version else row[3],
                                '合约': instrument,
                                '开仓日期': pd.to_datetime(row[7]) if is_new_version else pd.to_datetime(row[5]),
                                '买/卖': b_s,
                                '手数': lots,
                                '开仓价': open_price,
                                '昨结算': float(row[12]) if is_new_version else float(row[9]),
                                '成交价': transaction_price,
                                '平仓盈亏': float(row[14]) if is_new_version else float(row[11]),
                                '权利金收支': float(row[15]) if is_new_version else float(row[12]),
                                '交易盈亏': price_margin * lots * TRADING_UNIT[
                                    re.findall(r'[A-Za-z]*[^0-9-]+', instrument)[0].lower()],
                                '盈亏率': price_margin / open_price}
                        df = pd.DataFrame([data])
                        position_closed = pd.concat([position_closed, df], ignore_index=True)
                        position_closed['持仓天数'] = position_closed['平仓日期'] - position_closed['开仓日期']
                        position_closed['持仓天数'].apply(lambda x: x.days)
    print(f'\n{datetime.now()} | 信息 | 已提取所有结算单数据')

    return client_id, account, transaction_record, position_closed


# ---------------------------------------------------- 数据提取 结束 ----------------------------------------------------


# ---------------------------------------------------- 数据统计 开始 ----------------------------------------------------
def calculate_yield(net_worth: pd.Series, method: Literal['log', 'pct'] = STATISTIC_METHOD) -> pd.Series:
    """
    计算收益率
    calculate yield

    Args:
        net_worth:
        method:

    Returns:

    """
    benchmark = net_worth.shift(1).fillna(method='ffill')
    if method == 'pct':
        return (net_worth / benchmark - 1).astype('float64')  # 百分比变动法计算
    elif method == 'log':
        return np.log(net_worth) - np.log(benchmark)  # 对数变动法计算
    else:
        raise ValueError(f"检测到{method}, 必须为 'pct'|'log' 之一")


def convert_yield(net_yield: Union[pd.Series, float], to: Literal['log', 'pct'] = STATISTIC_METHOD) \
        -> Union[pd.Series, float]:
    """
    在对数收益率和百分比收益率之间进行转换\n
    convert yield between log return and percentage return\n

    Args:
        net_yield:
        to:

    Returns:
         (pd.Series): yield
    """
    if to == 'log':
        return np.log(net_yield + 1.0)
    elif to == 'pct':
        return np.exp(net_yield) - 1.0
    else:
        raise ValueError(f"检测到{to}, 必须为 'pct'|'log' 之一")


def merge_with_benchmark(data, reference, method: Literal['log', 'pct'] = STATISTIC_METHOD, how='inner'):
    data_temp = data
    data_temp['收益率'] = calculate_yield(data, method=method).dropna()
    reference_temp = reference
    reference_temp['收益率'] = calculate_yield(reference, method=method).dropna()
    return pd.merge(data_temp, reference_temp, on='日期', how=how)

def net_worth_calc(account: pd.DataFrame, method: Literal['log', 'pct'] = STATISTIC_METHOD):
    print(f'{datetime.now()} | 信息 | 开始净值化处理')
    net_worth = pd.DataFrame(columns=['日期', '净权益', '净值', '收益率', '份额', '份额变动'])
    net_worth['日期'] = account['日期']
    df = pd.merge(account, net_worth, on=['日期'])
    df['净权益'] = df['客户权益']
    df.loc[0, '份额'] = (df.iloc[0]['期初结存'] + df.iloc[0]['出入金']) / 1.0
    df.loc[0, '份额变动'] = df.iloc[0]['份额']
    df.loc[0, '净值'] = df.iloc[0]['净权益'] / df.iloc[0]['份额']
    for index in range(1, len(df.index)):
        if df.iloc[index]['出入金'] != 0:
            df.loc[index, '净值'] = (df.iloc[index]['净权益'] - df.iloc[index]['出入金']) / df.iloc[index - 1]['份额']
            df.loc[index, '份额变动'] = df.iloc[index]['出入金'] / df.iloc[index]['净值']
            df.loc[index, '份额'] = df.iloc[index - 1]['份额'] + df.iloc[index]['份额变动']
        else:
            df.loc[index, '净值'] = df.iloc[index]['净权益'] / df.iloc[index - 1]['份额']
            df.loc[index, '份额'] = df.iloc[index - 1]['份额']
    df['收益率'] = calculate_yield(df['净值'].astype('float64'), method=method)
    df.loc[0, '收益率'] = np.log(df.iloc[0]['净值']) if method == 'log' else df.iloc[0]['净值'] - 1.0
    df['份额变动'].fillna(0, inplace=True)
    net_worth = pd.DataFrame(df, columns=net_worth.columns)
    print(f'{datetime.now()} | 信息 | 已完成净值化处理')
    return net_worth


def annual_attribution_statistic(net_worth: pd.DataFrame, benchmark: Optional[pd.DataFrame] = None,
                                 method: Literal['log', 'pct'] = STATISTIC_METHOD):
    def accumulated_return(data: pd.Series):
        """
        累积收益率

        Args:
            data:

        Returns:

        """
        if data.name == '净值':
            return (data.iloc[-1] - data.iloc[0]) / data.iloc[0]
        elif data.name == '收益率':
            if method == 'log':
                return np.sum(data)
            elif method == 'pct':
                return np.prod(data + 1.0) - 1.0
        else:
            raise ValueError

    def annual_return(data: pd.Series):
        """
        年化收益率

        Args:
            data:

        Returns:

        """
        return np.power(1 + accumulated_return(data), 252 / len(data)) - 1

    def historical_volatility(data: pd.Series):
        """
        波动率
        收益率的标准差

        Args:
            data:

        Returns:

        """
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        rete = np.average(ret)
        return np.sqrt(np.sum(np.square(ret - rete)) / (len(ret) - 1))  # 等价ret.std(ddof=1)

    def annual_volatility(data: pd.Series):
        """
        年化波动率

        Args:
            data:

        Returns:

        """
        return historical_volatility(data) * np.sqrt(252)

    def downward_risk(data: pd.Series):
        """
        下行风险
        负收益率的标准差

        Args:
            data:

        Returns:

        """
        rfd = np.power(1 + RISK_FREE_INTEREST_RATE, 1 / 252) - 1
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        dif = ret - rfd
        negative_dif = dif[dif < 0]
        return np.sqrt(np.sum(np.square(negative_dif)) / (len(ret) - 1))

    def annual_downward_risk(data: pd.Series):
        """
        年化下行风险

        Args:
            data:

        Returns:

        """
        return downward_risk(data) * np.sqrt(252)

    def max_drawdown(data: pd.Series):
        """
        最大回撤

        Args:
            data:

        Returns:

        """
        if data.name == '净值':
            maximums = np.maximum.accumulate(data)
            return np.max(1 - data / maximums)
        elif data.name == '收益率':
            if method == 'log':
                cum_yield = np.cumsum(data)
                maximums = np.maximum.accumulate(cum_yield)
                return np.max(maximums - cum_yield)
            elif method == 'pct':
                cum_yield = np.cumprod(data + 1.0)
                maximums = np.maximum.accumulate(cum_yield)
                return np.max(1 - data / maximums)

    def historical_sharpe_ratio(data: pd.Series):
        """
        夏普比率
        等于(策略平均收益率 - 无风险利率)/区间波动率, 代表资产组合承担单位风险获得的回报情况
        若为正值, 代表可以获得的回报率高过波动风险; 若为负值, 代表操作风险大过于资产回报率
        数值越大, 表现越好
        [适用范围]: 当投资组合内的资产皆为风险性资产且分布为正态分布

        Args:
            data:

        Returns:

        """
        rfd = np.power(1 + RISK_FREE_INTEREST_RATE, 1 / 252) - 1
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        rete = np.average(ret)
        return (rete - rfd) / historical_volatility(data)

    def annual_sharpe_ratio(data: pd.Series):
        """
        年化夏普比率

        Args:
            data:

        Returns:

        """
        return (annual_return(data) - RISK_FREE_INTEREST_RATE) / annual_volatility(data)

    def historical_sortino_ratio(data: pd.Series):
        """
        索提诺比率
        与夏普比率类似, 不同的是它区分了波动的好坏, 在计算波动率时它所采用的不是标准差, 而是下行标准差
        这其中的隐含条件是投资组合的上涨（正回报率）符合投资人的需求, 不应计入风险调整
        数值越大, 业绩表现越好

        Args:
            data:

        Returns:

        """
        rfd = np.power(1 + RISK_FREE_INTEREST_RATE, 1 / 252) - 1
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        rete = np.average(ret)
        return (rete - rfd) / downward_risk(data)

    def annual_sortino_ratio(data: pd.Series):
        """
        年化索提诺比率

        Args:
            data:

        Returns:

        """
        return (annual_return(data) - RISK_FREE_INTEREST_RATE) / annual_downward_risk(data)

    def calmar_ratio(data: pd.Series):
        """
        卡玛比率
        等于(平均收益率-无风险利率)/最大回撤. 表明在单位回撤下获得的收益状况
        数值越大, 表现越好; 反之, 数值越小, 表现越差

        Args:
            data:

        Returns:

        """
        rfd = np.power(1 + RISK_FREE_INTEREST_RATE, 1 / 252) - 1
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        rete = np.average(ret)
        return (rete - rfd) / max_drawdown(data)

    def tail_ratio(data: pd.Series):
        """
        可以理解成衡量95分位收益与5分位的亏损的收益表现指标
        数值越大, 表现越好
        例如: tail ratio = 0.25, 5分位的亏损是95分位收益的四倍. 这样的策略在发生大额亏损的情况下很难在短时间内恢复
        使用范围: 均值回归策略, 这类型策略的最大风险在于左侧的尾部风险, 单次的大额回撤需要很长的时间才能够恢复

        Args:
            data:

        Returns:

        """
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        return np.percentile(ret, 0.95) / np.percentile(ret, 0.05)

    def gain_to_pain_ratio(data: pd.Series):
        """
        用于衡量资产组合正回报率总和与负回报率总和之间的比率
        数值越大, 表现越好

        Args:
            data:

        Returns:

        """
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        return np.abs(np.sum(ret[ret > 0]) / np.sum(ret[ret < 0]))

    def common_sense_ratio(data: pd.Series):
        """
        数值大于1: 策略盈利
        数值小于1: 策略亏损.
        使用范围: 均值回归策略, 趋势追踪策略.

        Args:
            data:

        Returns:

        """
        return tail_ratio(data) * gain_to_pain_ratio(data)

    def skewness(data: pd.Series):
        """
        偏度
        标准正态分布偏度为0, 大于0表示收益分布与正态分布相比为正偏或右偏
        正偏(): 算数平均数 > 中位数 > 众数
        负偏(): 众数 > 中位数 > 算数平均数

        Args:
            data:

        Returns:

        """
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        rete = np.average(ret)
        return np.mean(np.power(ret - rete, 3)) / np.power(np.std(ret, ddof=1), 3)

    def kurtosis(data: pd.Series):
        """
        峰度
        标准正态分布峰度接近0, 大于0表示收益的分布与正态分布相比较为陡峭
        常峰度正态分布峰度等于3, 超额态分布峰度等于-3, 正峰度(尖峰)大于3, 负峰度(扁峰)小于3

        Args:
            data:

        Returns:

        """
        ret = data  # 收益率序列
        if data.name == "净值":
            ret = calculate_yield(data, method=method).dropna()
        rete = np.average(ret)
        return np.mean(np.power(ret - rete, 4)) / np.power(np.mean(np.power(ret - rete, 2)), 2) - 3

    def alpha(data: pd.DataFrame, reference: pd.DataFrame):
        """
        alpha
        反映了资产组合获得的与基准波动无关的超额回报
        数值越大, 表现越好

        Args:
            data:
            reference:

        Returns:

        """
        combined_temp = merge_with_benchmark(data, reference, method=method, how='inner')
        ret = np.prod(combined_temp['收益率_x'] + 1) - 1
        remt = np.prod(combined_temp['收益率_y'] + 1) - 1
        rft = np.power(1 + RISK_FREE_INTEREST_RATE, len(combined_temp) / 252) - 1
        return ret - (rft + beta(data, reference) * (remt - rft))

    def beta(data: pd.DataFrame, reference: pd.DataFrame):
        """
        beta
        用以度量某项资产或资产组合相对基准的波动性
        数值大于1: 资产组合的波动性比基准波动大
        数值等于1: 资产组合与基准同步变化
        数值大于0小于1: 资产组合的波动性比基准小

        Args:
            data:
            reference:

        Returns:

        """
        combined_temp = merge_with_benchmark(data, reference, method=method, how='inner')
        return np.cov(combined_temp['收益率_x'], combined_temp['收益率_y'])[0][1] / np.var(combined_temp['收益率_y'], ddof=1)

    def treynor_ratio(data: pd.DataFrame, reference: pd.DataFrame):
        """
        特雷诺比率
        是以基金收益的系统风险作为基金绩效调整的因子, 反映基金承担单位系统风险所获得的超额收益
        数值越大, 承担单位系统风险所获得的超额收益越高

        Args:
            data:
            reference:

        Returns:

        """
        rfd = np.power(1 + RISK_FREE_INTEREST_RATE, 1 / 252) - 1
        try:
            ret = data['收益率']  # 收益率序列
        except KeyError:
            ret = calculate_yield(data['净值'], method=method).dropna()
        rete = np.average(ret)
        return (rete - rfd) / beta(data, reference)

    def stability_of_time_series(data: pd.DataFrame, reference: pd.DataFrame):
        """
        累计对数收益对时间t的回归的R^2, 也称为决定系数
        用于衡量基金与基准之间的变动同步程度
        数值大小反映了趋势线的估计值与对应的实际数据之间的拟合程度, 拟合程度越高, 趋势线的可靠性就越高
        R平方值是取值范围在0～1之间的数值, 数值越高, 变动同步程度越一致

        Args:
            data:
            reference:

        Returns:

        """
        combined_temp = merge_with_benchmark(data, reference, method=method, how='inner')
        x = combined_temp['收益率_x']
        y = combined_temp['收益率_y']
        zx = (x - np.mean(x)) / np.std(x, ddof=1)
        zy = (y - np.mean(y)) / np.std(y, ddof=1)
        r = np.sum((zx * zy) / (len(x) - 1))
        return np.square(r)

    def information_ratio(data: pd.DataFrame, reference: pd.DataFrame):
        """
        信息比率
        以马克维茨的均异模型为基础, 可以衡量基金的均异特性, 它表示单位主动风险所带来的超额收益
        数值越大, 表现越好

        Args:
            data:
            reference:

        Returns:

        """
        combined_temp = merge_with_benchmark(data, reference, method=method, how='inner')
        combined_temp['收益率_dif'] = combined_temp['收益率_x'] - combined_temp['收益率_y']
        return np.mean(combined_temp['收益率_dif']) / np.sqrt(np.var(combined_temp['收益率_dif'], ddof=1))

    def modigliani_ratio(data: pd.DataFrame, reference: pd.DataFrame):
        """
        夏普比率和信息比率的组合, 简称m2.
        资产组合相对于市场无风险利率的超额收益与基准相对于市场无风险利率的超额收益之间的差异,
        用于衡量资产组合相对于基准的超额收益.
        数值越大, 表现越好

        Args:
            data:
            reference:

        Returns:

        """
        combined_temp = merge_with_benchmark(data, reference, method=method, how='inner')
        rfd = np.power(1 + RISK_FREE_INTEREST_RATE, 1 / 252) - 1
        combined_temp['data_yield_dif_to_rf'] = combined_temp['收益率_x'] - rfd
        combined_temp['benchmark_yield_dif_to_rf'] = combined_temp['收益率_y'] - rfd
        rft = np.power(1 + RISK_FREE_INTEREST_RATE, len(combined_temp) / 252) - 1
        return np.mean(combined_temp['data_yield_dif_to_rf']) * np.sqrt(
            np.var(combined_temp['benchmark_yield_dif_to_rf'], ddof=1)) / np.sqrt(
            np.var(combined_temp['data_yield_dif_to_rf'], ddof=1)) + rft

    print(f'{datetime.now()} | 信息 | 开始年度归因分析')
    temp_net_worth = net_worth.set_index('日期')
    annual_statistic = pd.DataFrame(
        columns=['年度', '累计收益率', '年化收益率', '历史波动率', '年化波动率', '最大回撤', '下行风险', '年化下行风险', '最新夏普比率',
                 '年化夏普比率', '最新索提诺比率', '年化索提诺比率', '卡玛比率', '盈亏率', 'Tail比率', 'CommonSense比率', '偏度',
                 '峰度', ])
    annual_statistic['年度'] = range(temp_net_worth.index[0].year, temp_net_worth.index[-1].year + 1)
    annual_statistic.set_index('年度', inplace=True)
    for year in annual_statistic.index:
        data_set = temp_net_worth.loc[f'{year}-01-01': f'{year}-12-31']
        annual_statistic.loc[year, '累计收益率'] = accumulated_return(data_set['收益率'])
        annual_statistic.loc[year, '年化收益率'] = annual_return(data_set['收益率'])
        annual_statistic.loc[year, '历史波动率'] = historical_volatility(data_set['收益率'])
        annual_statistic.loc[year, '年化波动率'] = annual_volatility(data_set['收益率'])
        annual_statistic.loc[year, '最大回撤'] = max_drawdown(data_set['净值'])  # 这里用对数收益率可能导致回撤大于100%, 可以转化但没必要
        annual_statistic.loc[year, '下行风险'] = downward_risk(data_set['收益率'])
        annual_statistic.loc[year, '年化下行风险'] = annual_downward_risk(data_set['收益率'])
        annual_statistic.loc[year, '最新夏普比率'] = historical_sharpe_ratio(data_set['收益率'])
        annual_statistic.loc[year, '年化夏普比率'] = annual_sharpe_ratio(data_set['收益率'])
        annual_statistic.loc[year, '最新索提诺比率'] = historical_sortino_ratio(data_set['收益率'])
        annual_statistic.loc[year, '年化索提诺比率'] = annual_sortino_ratio(data_set['收益率'])
        annual_statistic.loc[year, '卡玛比率'] = calmar_ratio(data_set['收益率'])
        annual_statistic.loc[year, '盈亏率'] = gain_to_pain_ratio(data_set['收益率'])
        annual_statistic.loc[year, 'Tail比率'] = tail_ratio(data_set['收益率'])
        annual_statistic.loc[year, 'CommonSense比率'] = common_sense_ratio(data_set['收益率'])
        annual_statistic.loc[year, '偏度'] = skewness(data_set['收益率'])
        annual_statistic.loc[year, '峰度'] = kurtosis(data_set['收益率'])
    if benchmark is not None:
        for year in annual_statistic.index:
            data_set = temp_net_worth.loc[f'{year}-01-01': f'{year}-12-31']
            annual_statistic.loc[year, 'Alpha'] = alpha(data_set, benchmark)
            annual_statistic.loc[year, 'Beta'] = beta(data_set, benchmark)
            annual_statistic.loc[year, '特雷诺比率'] = treynor_ratio(data_set, benchmark)
            annual_statistic.loc[year, 'R2'] = stability_of_time_series(data_set, benchmark)
            annual_statistic.loc[year, '信息比率'] = information_ratio(infodata_set, benchmark)
            annual_statistic.loc[year, 'M2'] = modigliani_ratio(data_set, benchmark)
    annual_statistic.reset_index(inplace=True)
    print(f'{datetime.now()} | 信息 | 已完成年度归因分析')
    return annual_statistic


def data_statistic(transaction_record, position_closed):
    print(f'{datetime.now()} | 信息 | 开始交易数据统计')
    statistic_by_contracts = pd.DataFrame(
        columns=['品种', '合约', '平仓盈亏', '净利润', '交易次数', '交易手数', '盈利次数',
                 '盈利手数', '交易成功率', '交易盈亏率', '均次盈亏', '均手盈亏', '最大盈利',
                 '最大亏损', '成交额'])
    position_closed_group_by_contracts = position_closed.groupby('合约')
    statistic_by_contracts['合约'] = position_closed_group_by_contracts.groups.keys()
    statistic_by_contracts = statistic_by_contracts.set_index('合约')
    for index in statistic_by_contracts.index:
        matched_string = re.findall(r'[A-Za-z]*[^0-9-]+', index)
        if len(matched_string) == 2:
            statistic_by_contracts.loc[index][
                '品种'] = f"{CONTRACT_CODE[matched_string[0].lower()]}{'看涨' if matched_string[1].lower() == 'c' else '看跌'}期权"
        else:
            statistic_by_contracts.loc[index]['品种'] = CONTRACT_CODE[matched_string[0].lower()]
    statistic_by_contracts['平仓盈亏'] = position_closed_group_by_contracts['交易盈亏'].sum().astype('float64')
    statistic_by_contracts['净利润'] = position_closed_group_by_contracts['交易盈亏'].sum().astype('float64') - \
                                       transaction_record.groupby('合约')['手续费'].sum().astype('float64')
    statistic_by_contracts['交易次数'] = position_closed_group_by_contracts['品种'].count().astype('int64')
    statistic_by_contracts['交易手数'] = position_closed_group_by_contracts['手数'].sum().astype('int64')
    statistic_by_contracts['盈利次数'] = position_closed_group_by_contracts.apply(lambda x: sum(x['交易盈亏'] > 0))
    statistic_by_contracts['盈利手数'] = position_closed_group_by_contracts.apply(
        lambda x: sum(x[x['交易盈亏'] > 0]['手数']))
    statistic_by_contracts['交易成功率'] = round(
        statistic_by_contracts['盈利次数'] / statistic_by_contracts['交易次数'], 4)
    statistic_by_contracts['交易盈亏率'] = round(
        statistic_by_contracts['盈利手数'] / statistic_by_contracts['交易手数'], 4)
    statistic_by_contracts['均次盈亏'] = round(statistic_by_contracts['平仓盈亏'] / statistic_by_contracts['交易次数'],
                                               2)
    statistic_by_contracts['均手盈亏'] = round(statistic_by_contracts['平仓盈亏'] / statistic_by_contracts['交易手数'],
                                               2)
    statistic_by_contracts['最大盈利'] = position_closed_group_by_contracts.apply(
        lambda x: max(max(x['交易盈亏']), 0.)).astype('float64')
    statistic_by_contracts['最大亏损'] = position_closed_group_by_contracts.apply(
        lambda x: min(min(x['交易盈亏']), 0.)).astype('float64')
    statistic_by_contracts['成交额'] = transaction_record.groupby('合约')['成交额'].sum().astype('float64')
    statistic_by_contracts = statistic_by_contracts.reset_index()

    statistic_by_categories = pd.DataFrame(
        columns=['品种', '平仓盈亏', '净利润', '交易次数', '交易手数', '盈利次数', '盈利手数',
                 '交易成功率', '交易盈亏率', '均次盈亏', '均手盈亏', '最大盈利', '最大亏损',
                 '成交额'])
    contracts_analysis_group_by_categories = statistic_by_contracts.groupby('品种')
    statistic_by_categories['品种'] = contracts_analysis_group_by_categories.groups.keys()
    statistic_by_categories = statistic_by_categories.set_index('品种')
    statistic_by_categories['平仓盈亏'] = contracts_analysis_group_by_categories['平仓盈亏'].sum()
    statistic_by_categories['净利润'] = contracts_analysis_group_by_categories['净利润'].sum()
    statistic_by_categories['交易次数'] = contracts_analysis_group_by_categories['交易次数'].sum()
    statistic_by_categories['交易手数'] = contracts_analysis_group_by_categories['交易手数'].sum()
    statistic_by_categories['盈利次数'] = contracts_analysis_group_by_categories['盈利次数'].sum()
    statistic_by_categories['盈利手数'] = contracts_analysis_group_by_categories['盈利手数'].sum()
    statistic_by_categories['交易成功率'] = round(
        statistic_by_categories['盈利次数'] / statistic_by_categories['交易次数'], 4)
    statistic_by_categories['交易盈亏率'] = round(
        statistic_by_categories['盈利手数'] / statistic_by_categories['交易手数'], 4)
    statistic_by_categories['均次盈亏'] = round(
        statistic_by_categories['平仓盈亏'] / statistic_by_categories['交易次数'], 2)
    statistic_by_categories['均手盈亏'] = round(
        statistic_by_categories['平仓盈亏'] / statistic_by_categories['交易手数'], 2)
    statistic_by_categories['最大盈利'] = contracts_analysis_group_by_categories.apply(
        lambda x: max(max(x['最大盈利']), 0.)).astype('float64')
    statistic_by_categories['最大亏损'] = contracts_analysis_group_by_categories.apply(
        lambda x: min(min(x['最大亏损']), 0.)).astype('float64')
    statistic_by_categories['成交额'] = contracts_analysis_group_by_categories['成交额'].sum()
    statistic_by_categories = statistic_by_categories.reset_index()

    statistic_by_trade_direction = pd.DataFrame(
        columns=['统计指标', '总盈利', '总亏损', '总盈利/总亏损', '手续费', '净利润',
                 '盈利比率', '盈利手数', '亏损手数', '持平手数', '平均盈利', '平均亏损',
                 '平均盈利/平均亏损', '平均手续费', '平均净利润', '最大盈利', '最大亏损',
                 '最大盈利/总盈利', '最大亏损/总亏损', '净利润/最大亏损'])
    position_closed_group_by_trade_direction = position_closed.groupby('买/卖')
    statistic_by_trade_direction['统计指标'] = position_closed_group_by_trade_direction.groups.keys()
    statistic_by_trade_direction = statistic_by_trade_direction.set_index('统计指标')
    statistic_by_trade_direction['总盈利'] = position_closed_group_by_trade_direction.apply(
        lambda x: sum(x[x['交易盈亏'] > 0]['交易盈亏']))
    statistic_by_trade_direction['总亏损'] = position_closed_group_by_trade_direction.apply(
        lambda x: sum(x[x['交易盈亏'] < 0]['交易盈亏']))
    statistic_by_trade_direction['总盈利/总亏损'] = round(
        abs(statistic_by_trade_direction['总盈利'] / statistic_by_trade_direction['总亏损']), 4)
    statistic_by_trade_direction['手续费'] = transaction_record.groupby('买/卖')['手续费'].sum().astype('float64')
    statistic_by_trade_direction['净利润'] = statistic_by_trade_direction['总盈利'] + statistic_by_trade_direction[
        '总亏损'] - transaction_record.groupby('买/卖')['手续费'].sum().astype('float64')
    statistic_by_trade_direction['盈利手数'] = position_closed_group_by_trade_direction.apply(
        lambda x: sum(x[x['交易盈亏'] > 0]['手数']))
    statistic_by_trade_direction['亏损手数'] = position_closed_group_by_trade_direction.apply(
        lambda x: sum(x[x['交易盈亏'] < 0]['手数']))
    statistic_by_trade_direction['持平手数'] = position_closed_group_by_trade_direction.apply(
        lambda x: sum(x[x['交易盈亏'] == 0]['手数']))
    statistic_by_trade_direction['盈利比率'] = round(statistic_by_trade_direction['盈利手数'] / (
            statistic_by_trade_direction['盈利手数'] + statistic_by_trade_direction['亏损手数'] +
            statistic_by_trade_direction['持平手数']), 4)
    statistic_by_trade_direction['平均盈利'] = round(
        statistic_by_trade_direction['总盈利'] / statistic_by_trade_direction['盈利手数'], 2)
    statistic_by_trade_direction['平均亏损'] = round(
        statistic_by_trade_direction['总亏损'] / statistic_by_trade_direction['亏损手数'], 2)
    statistic_by_trade_direction['平均盈利/平均亏损'] = round(
        abs(statistic_by_trade_direction['平均盈利'] / statistic_by_trade_direction['平均亏损']), 4)
    statistic_by_trade_direction['平均手续费'] = round(statistic_by_trade_direction['手续费'] / (
            statistic_by_trade_direction['盈利手数'] + statistic_by_trade_direction['亏损手数'] +
            statistic_by_trade_direction['持平手数']), 2)
    statistic_by_trade_direction['平均净利润'] = statistic_by_trade_direction['平均盈利'] + \
                                                 statistic_by_trade_direction['平均亏损'] - \
                                                 statistic_by_trade_direction['平均手续费']
    statistic_by_trade_direction['最大盈利'] = position_closed_group_by_trade_direction['交易盈亏'].max()
    statistic_by_trade_direction['最大亏损'] = position_closed_group_by_trade_direction['交易盈亏'].min()
    statistic_by_trade_direction['最大盈利/总盈利'] = round(
        statistic_by_trade_direction['最大盈利'] / statistic_by_trade_direction['总盈利'], 4)
    statistic_by_trade_direction['最大亏损/总亏损'] = round(
        statistic_by_trade_direction['最大亏损'] / statistic_by_trade_direction['总亏损'], 4)
    statistic_by_trade_direction['净利润/最大亏损'] = round(
        abs(statistic_by_trade_direction['净利润'] / statistic_by_trade_direction['最大亏损']), 4)
    statistic_by_trade_direction = statistic_by_trade_direction.reset_index()
    print(f'{datetime.now()} | 信息 | 已完成交易数据统计')
    return statistic_by_contracts, statistic_by_categories, statistic_by_trade_direction


# ---------------------------------------------------- 数据统计 结束 ----------------------------------------------------


# --------------------------------------------------- 数据格式化 开始 ---------------------------------------------------
def excel_data_format(excel_file):
    def cell_format_by_columns(worksheet):
        """
        单元格根据列表头格式化
        :param worksheet:
        :return:
        """
        for column_dataset in worksheet.columns:
            header = column_dataset[0].value
            for row_index in range(worksheet.min_row, worksheet.max_row):
                if '日期' in header:
                    column_dataset[row_index].number_format = numbers.FORMAT_DATE_YYYYMMDD2
                elif any([c in header for c in ['风险度', '率', '/']]):
                    column_dataset[row_index].number_format = numbers.FORMAT_PERCENTAGE_00
                elif any([c in header for c in ['次数', '手数']]):
                    column_dataset[row_index].number_format = numbers.FORMAT_NUMBER
                elif any([c in header for c in ['利润', '盈亏', '盈利', '亏损', '结存', '权益', '保证金', '出入金', '成交额']]):
                    column_dataset[row_index].number_format = numbers.BUILTIN_FORMATS[39]

    def cell_format_by_rows(worksheet):
        """
        单元格根据行表头格式化
        :param worksheet:
        :return:
        """
        for row_dataset in worksheet.rows:
            header = row_dataset[0].value
            for column_index in range(worksheet.min_column, worksheet.max_column):
                if '日期' in header:
                    row_dataset[column_index].number_format = numbers.FORMAT_DATE_YYYYMMDD2
                elif any([c in header for c in ['风险度', '率', '/']]):
                    row_dataset[column_index].number_format = numbers.FORMAT_PERCENTAGE_00
                elif any([c in header for c in ['次数', '手数']]):
                    row_dataset[column_index].number_format = numbers.FORMAT_NUMBER
                elif any([c in header for c in ['利润', '盈亏', '盈利', '亏损', '结存', '权益', '保证金', '出入金', '成交额']]):
                    row_dataset[column_index].number_format = numbers.BUILTIN_FORMATS[39]

    def dimension_format(worksheet, dimension='columns'):
        """
        行列格式化
        :param worksheet:
        :param dimension:
        :return:
        """
        if dimension == 'columns':
            for column_index in get_column_interval(worksheet.min_column, worksheet.max_column):
                worksheet.column_dimensions[column_index].auto_size = True
                worksheet.column_dimensions[column_index].best_fit = True
        elif dimension == 'rows':
            for row_index in range(worksheet.min_row, worksheet.max_row):
                pass

    def data_transposition(workbook, sheet_name, index=None):
        """
        数据转置
        :param workbook:
        :param sheet_name:
        :param index:
        :return:
        """
        data = workbook[sheet_name].values
        if index:
            cols = next(data)[1:]
            data = list(data)
            index = [r[0] for r in data]
            data = (islice(r, 1, None) for r in data)
        else:
            cols = next(data)[0:]
            data = list(data)
            data = (islice(r, 0, None) for r in data)
        df = pd.DataFrame(data, index=index, columns=cols).T
        df.reset_index(level=0, inplace=True)
        idx = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook.worksheets[idx])
        workbook.create_sheet(sheet_name, idx)
        for r in dataframe_to_rows(df, index=True if index else False, header=False):
            workbook[sheet_name].append(r)
        for cell in workbook[sheet_name]['A']:
            cell.style = 'Pandas'

    wb = load_workbook(excel_file)
    print(f'{datetime.now()} | 信息 | 开始Excel数据格式化')
    for ws_name in wb.sheetnames:
        cell_format_by_columns(wb[ws_name])
        dimension_format(wb[ws_name])
        if ws_name == '交易分析(按买卖)':
            data_transposition(wb, '交易分析(按买卖)')
            cell_format_by_rows(wb[ws_name])
            dimension_format(wb[ws_name])
    wb.save(excel_file)
    wb.close()
    print(f'{datetime.now()} | 信息 | 已生成Excel数据表')


# --------------------------------------------------- 数据格式化 结束 ---------------------------------------------------


# ---------------------------------------------------- 生成图表 开始 ----------------------------------------------------
def excel_create_chart(excel_file):
    """
    生成图表
    :param excel_file:
    :return:
    """
    wb = load_workbook(excel_file)
    print(f'{datetime.now()} | 信息 | 开始Excel图表渲染')
    # 进行净值走势图渲染
    net_worth_sheet = wb['账户净值']
    net_worth_chart_sheet = wb.create_chartsheet(title='净值走势图')
    chart1 = LineChart()
    dates1 = Reference(net_worth_sheet, min_col=1, min_row=2, max_row=len(net_worth_sheet['A']))
    data1 = Reference(net_worth_sheet, min_col=3, min_row=1, max_row=len(net_worth_sheet['C']))
    chart1.add_data(data1, titles_from_data=True)
    chart1.y_axis = NumericAxis(title='净值', majorTickMark='out')
    chart1.x_axis = TextAxis(majorTickMark='out', tickLblSkip=10, tickMarkSkip=10, noMultiLvlLbl=True,
                             numFmt='yyyy-mm-dd')
    chart1.legend = None
    chart1.title = str()
    chart1.style = 1
    chart1.set_categories(dates1)
    net_worth_chart_sheet.add_chart(chart1)
    # 进行权益走势图渲染
    account_sheet = wb['账户统计']
    account_chart_sheet = wb.create_chartsheet(title='权益走势图')
    dates2 = Reference(account_sheet, min_col=1, min_row=2, max_row=len(account_sheet['A']))
    chart2 = LineChart(varyColors=True)
    data2 = Reference(account_sheet, min_col=8, min_row=1, max_row=len(account_sheet['H']))
    chart2.add_data(data2, titles_from_data=True)
    chart2.y_axis = NumericAxis(title='权益', majorTickMark='out')
    chart2.legend = None
    chart2.set_categories(dates2)
    chart3 = BarChart()
    data3 = Reference(account_sheet, min_col=10, min_row=1, max_row=len(account_sheet['J']))
    chart3.add_data(data3, titles_from_data=True)
    chart3.y_axis = NumericAxis(axId=200, title='风险度', majorGridlines=None, majorTickMark='out', crosses='max')
    chart3.x_axis = TextAxis(majorTickMark='out', tickLblSkip=10, tickMarkSkip=10, noMultiLvlLbl=True,
                             numFmt='yyyy-mm-dd')
    chart3.legend = None
    chart3.set_categories(dates2)
    chart2 += chart3
    account_chart_sheet.add_chart(chart2)
    # 进行交易分布图的渲染
    categories_analysis_sheet = wb['交易分析(按品种)']
    trading_frequency_analysis_sheet = wb.create_chartsheet(title='交易分布图')
    labels = Reference(categories_analysis_sheet, min_col=1, min_row=2, max_row=len(categories_analysis_sheet['A']))
    chart4 = PieChart(varyColors=True)
    chart4.style = 34
    data4 = Reference(categories_analysis_sheet, min_col=4, min_row=2, max_row=len(categories_analysis_sheet['D']))
    chart4.add_data(data4)
    chart4.set_categories(labels)
    chart4.legend = None
    # chart4.series[0].data_points = [DataPoint(idx=i, explosion=8)
    #                                 for i in range(len(categories_analysis_sheet['D']) - 1)]
    chart4.series[0].dLbls = DataLabelList(dLblPos='bestFit', showPercent=True, showCatName=True, showVal=True,
                                           showLeaderLines=True)
    chart4.layout = Layout(manualLayout=ManualLayout(x=0, y=0, h=0.75, w=0.75, xMode='factor', yMode='factor'))
    trading_frequency_analysis_sheet.add_chart(chart4)
    # 进行品种盈亏图的渲染
    categories_win_and_loss_chart_sheet = wb.create_chartsheet(title='品种盈亏图')
    chart5 = BarChart(barDir='col')
    chart5.style = 18
    data5 = Reference(categories_analysis_sheet, min_col=2, min_row=2, max_row=len(categories_analysis_sheet['B']))
    chart5.add_data(data5)
    chart5.set_categories(labels)
    chart5.legend = None
    chart5.series[0].dLbls = DataLabelList(showVal=True)
    chart5.y_axis = NumericAxis(title='平仓盈亏', majorTickMark='out', minorTickMark='out')
    categories_win_and_loss_chart_sheet.add_chart(chart5)
    # 进行交易盈亏图的渲染
    trading_win_and_loss_chart_sheet = wb.create_chartsheet(title='交易盈亏图')
    chart6 = RadarChart()
    chart6.style = 24
    data6 = Reference(categories_analysis_sheet, min_col=8, max_col=9, min_row=1,
                      max_row=categories_analysis_sheet.max_row)
    chart6.add_data(data6, titles_from_data=True)
    chart6.set_categories(labels)
    trading_win_and_loss_chart_sheet.add_chart(chart6)
    # 图表保存
    wb.save(excel_file)
    wb.close()
    # 输出信息
    print(f'{datetime.now()} | 信息 | 已生成Excel图表')


# ---------------------------------------------------- 生成图表 结束 ----------------------------------------------------


# -------------------------------------------------- 生成excel文件 开始 -------------------------------------------------
def output_excel(net_worth, annual_statistic, account, transaction_record, position_closed, contracts_analysis,
                 categories_analysis, trade_direction_analysis, client_id=''):
    try:
        with pd.ExcelWriter(os.path.join(BASE_DIR, client_id + '交易统计.xlsx'),
                            mode='w',
                            engine="openpyxl") as writer:
            net_worth.to_excel(writer, sheet_name='账户净值', index=False)
            annual_statistic.to_excel(writer, sheet_name='年度归因', index=False)
            account.to_excel(writer, sheet_name='账户统计', index=False)
            transaction_record.to_excel(writer, sheet_name='交易记录', index=False)
            position_closed.to_excel(writer, sheet_name='平仓明细', index=False)
            contracts_analysis.to_excel(writer, sheet_name='交易分析(按合约)', index=False)
            categories_analysis.to_excel(writer, sheet_name='交易分析(按品种)', index=False)
            trade_direction_analysis.to_excel(writer, sheet_name='交易分析(按买卖)', index=False)
        excel_data_format(os.path.join(BASE_DIR, client_id + '交易统计.xlsx'))
        excel_create_chart(os.path.join(BASE_DIR, client_id + '交易统计.xlsx'))
        input(f'{datetime.now()} | 信息 | 任务结束, 感谢您的使用, 按任意键退出!\n')
        # raise SystemExit()
    except PermissionError:
        input(f'{datetime.now()} | 错误 | 分析结果写入Excel被拒绝, 请检查文件是否已打开, 按任意键退出!\n')
        # raise SystemExit()


# -------------------------------------------------- 生成excel文件 结束 -------------------------------------------------


# ---------------------------------------------------- 终端命令 开始 ----------------------------------------------------
def main(argv):
    client_id = ''
    files_folder = ''
    try:
        opts, args = getopt.getopt(argv, "hd:i:", ["dir=", "id="])
    except getopt.GetoptError:
        print(
            '参数选项:\n-d/--dir <settlement statement files\' folder  结算单文件夹路径>\n-i/--id <client id  客户号>')
        sys.exit(2)
    if len(opts) != 0:
        for opt, arg in opts:
            if opt == '-h':
                print(
                    '参数选项:\n-d/--dir <settlement statement files\' folder  结算单文件夹路径>\n-i/--id <client id  客户号>')
                sys.exit()
            elif opt in ("-d", "--dir"):
                files_folder = arg
            elif opt in ("-i", "--id"):
                client_id = arg
    statement_list = read_statement_files(files_folder)
    client_id, account, transaction_record, position_closed = data_extract(statement_list, client_id=client_id)
    net_worth = net_worth_calc(account)
    annual_statistic = annual_attribution_statistic(net_worth)
    contracts_analysis, categories_analysis, trade_direction_analysis = data_statistic(transaction_record,
                                                                                       position_closed)
    output_excel(net_worth, annual_statistic, account, transaction_record, position_closed, contracts_analysis,
                 categories_analysis, trade_direction_analysis, client_id=client_id)


if __name__ == '__main__':
    main(sys.argv[1:])
# ---------------------------------------------------- 终端命令 结束 ----------------------------------------------------
