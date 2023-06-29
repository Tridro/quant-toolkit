#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time         : 2023/5/30 10:53
# @Author       : Tridro
# @Email        : tridro@beneorigin.com
# @Project      : tools
# @File         : automatic_calculation_of_main_contracts_margin_requirement.py
# @Software     : PyCharm
# All Copyright Reserved

import os
import sys
import re
from math import isnan
from datetime import datetime, timedelta

import pandas as pd
from tqsdk import TqApi, TqAuth
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_interval

BASE_DIR = os.path.dirname(os.path.realpath(sys.argv[0]))
CURRENT_TIME = datetime.now()

os.system("mode con cols=71 lines=30")
sys.stdout.write(f"+{'-' * 68}+\n"
                 f"|{'自动计算主力合约保证金需求':^55}|\n"
                 f"+{'-' * 68}+\n"
                 f"|{' ':68}|\n"
                 f"|{'数据行情由天勤量化提供支持，如不同意天勤量化使用条款，请终止使用':^36}|\n"
                 f"|{' ':68}|\n"
                 f"+{'-' * 68}+\n\n")
tq_account = input('天勤量化账号: ')
tq_password = input('天勤量化密码: ')
try:
    api = TqApi(auth=TqAuth(tq_account, tq_password))
    print(f'\n{datetime.now()} | 信息 | 天勤API认证成功')
except Exception as err:
    input(f"\n{datetime.now()} | 错误 | {err.args[0]} 按任意键程序将退出, 请检查后重试\n")
    raise SystemExit()

margin_ratio_option_input = input('\n1: 交易所保证金(默认)\n2: 公司保证金\n请选择保证金比率(输入数字, 回车采用默认值): ') or 1
margin_ratio_option = '公司保证金' if margin_ratio_option_input == 2 else '交易所保证金'

if os.path.exists(os.path.join(BASE_DIR, 'margin_ratio.xlsx')):
    ratio = pd.read_excel(os.path.join(BASE_DIR, 'margin_ratio.xlsx'), index_col=1, engine='openpyxl')
    print(f'\n{datetime.now()} | 信息 | 已读取 margin_ratio.xlsx')
    ratio['交易所'].ffill(inplace=True)
else:
    input(f'\n{datetime.now()} | 错误 | {BASE_DIR} 目录下未找到 margin_ratio.xlsx, 按任意键程序将退出, 请检查后重试\n')
    raise SystemExit()

res = api.query_quotes(ins_class='CONT', expired=False)
quotes = api.get_quote_list(symbols=res)

underlying_symbols = [quote.underlying_symbol for quote in quotes]
contracts = [symbol.split('.')[1] for symbol in underlying_symbols]
contracts_category = [re.sub(r'\d+', '', contract).lower() for contract in contracts]


def get_trading_date(quote_datetime: datetime):
    if quote_datetime.hour in range(9, 16):
        if CURRENT_TIME.isoweekday() in range(1, 5) and CURRENT_TIME.hour in range(15, 21):
            return datetime(year=quote_datetime.year, month=quote_datetime.month, day=quote_datetime.day) + \
                timedelta(days=1)
        elif CURRENT_TIME.isoweekday() == 5 and CURRENT_TIME.hour in range(15, 21):
            return datetime(year=quote_datetime.year, month=quote_datetime.month, day=quote_datetime.day) + \
                timedelta(days=3)
        else:
            return datetime(year=quote_datetime.year, month=quote_datetime.month, day=quote_datetime.day)
    elif quote_datetime.hour in range(0, 3):
        t_day = datetime(year=quote_datetime.year, month=quote_datetime.month, day=quote_datetime.day)
        if t_day.isoweekday() == 6:
            return t_day + timedelta(days=2)
        else:
            return t_day
    elif quote_datetime.hour in range(20, 24):
        t_day = datetime(year=quote_datetime.year, month=quote_datetime.month, day=quote_datetime.day) + \
                timedelta(days=1)
        if t_day.isoweekday() == 5:
            return t_day + timedelta(days=3)
        else:
            return t_day


trade_datetime = {category: get_trading_date(datetime.fromisoformat(quote.datetime)) for category, quote in
                  zip(contracts_category, quotes)}
major_contracts = {category: contract for category, contract in zip(contracts_category, contracts)}
if CURRENT_TIME.isoweekday() in range(1, 6) and CURRENT_TIME.hour in range(15, 21):
    print(f'{datetime.now()} | 信息 | 当前时间为收盘时间, 将预估下一个交易日保证金 (采用当日交易均价, 股指保证金可能会不准确)')
    pre_settlement = {category: quote.average if not isnan(quote.average) else quote.pre_settlement
                      for category, quote in zip(contracts_category, quotes)}
else:
    print(f'{datetime.now()} | 信息 | 当前时间非收盘时间, 将计算当前交易日保证金')
    pre_settlement = {category: quote.pre_settlement for category, quote in zip(contracts_category, quotes)}

source_trade_datetime = pd.Series(trade_datetime, index=trade_datetime.keys())
source_pre_settlement = pd.Series(pre_settlement, index=pre_settlement.keys())
source_major_contracts = pd.Series(major_contracts, index=major_contracts.keys())

output = pd.DataFrame(columns=['交易所', '名称', '主力合约', '交易日期', '当日保证金', '昨日结算价', '保证金比率'])
output[['交易所', '名称']] = ratio[['交易所', '名称']]
output['交易日期'] = source_trade_datetime
output['主力合约'] = source_major_contracts
output['保证金比率'] = ratio[margin_ratio_option]
output['昨日结算价'] = source_pre_settlement
output['当日保证金'] = ratio['合约乘数'] * output['昨日结算价'] * output['保证金比率']
output.set_index(['交易所', '名称'], inplace=True)
if CURRENT_TIME.isoweekday() in range(1, 6) and CURRENT_TIME.hour in range(15, 21):
    output.rename(columns={'昨日结算价': '预估结算价', '当日保证金': '预估保证金'}, inplace=True)


def dimension_format(worksheet):
    """
    行列格式化
    :param worksheet:
    :return:
    """
    for column_index in get_column_interval(worksheet.min_column, worksheet.max_column):
        worksheet.column_dimensions[column_index].auto_size = True
        worksheet.column_dimensions[column_index].best_fit = True


def cell_format_by_columns(worksheet):
    """
    单元格根据列表头格式化
    :param worksheet:
    :return:
    """
    for column_dataset in worksheet.columns:
        header = column_dataset[0].value
        for row_index in range(worksheet.min_row, worksheet.max_row):
            if '日期' in header:
                column_dataset[row_index].number_format = numbers.FORMAT_DATE_YYYYMMDD2
            elif '率' in header or '/' in header or '%' in header:
                column_dataset[row_index].number_format = numbers.FORMAT_PERCENTAGE_00
            elif '保证金' in header or '结算价' in header:
                column_dataset[row_index].number_format = numbers.BUILTIN_FORMATS[39]


try:
    output_file_name = os.path.join(BASE_DIR, f"{datetime.now().strftime('%Y%m%d')}保证金.xlsx")
    output.to_excel(output_file_name, engine='openpyxl')
    print(f"{datetime.now()} | 信息 | 已生成 <{datetime.now().strftime('%Y%m%d')}保证金.xlsx>")
    wb = load_workbook(output_file_name)
    print(f'{datetime.now()} | 信息 | 开始Excel数据格式化')
    for ws_name in wb.sheetnames:
        cell_format_by_columns(wb[ws_name])
        dimension_format(wb[ws_name])
    wb.save(output_file_name)
    wb.close()
    print(f'{datetime.now()} | 信息 | 完成Excel数据格式化')
    input(f'{datetime.now()} | 信息 | 任务结束, 按任意键退出\n')
except PermissionError:
    input(f'{datetime.now()} | 错误 | 写入Excel被拒绝, 请检查文件是否已打开, 按任意键退出\n')
    raise SystemExit()
