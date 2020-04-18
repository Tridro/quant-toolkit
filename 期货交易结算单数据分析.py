#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2019/10/16 11:04
# @Author   : 周驰卓
# @Company  : 光大期货
# @Site     :
# @File     : 期货交易结算单数据分析.py
# @Software : PyCharm

import getopt
import os.path
import re
import sys
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.chart import Reference, LineChart, PieChart, BarChart, RadarChart
from openpyxl.chart.axis import NumericAxis
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout

# ---------------------------------------------------- 基础数据 开始 ----------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CONTRACT_CODE = {'if': '沪深300股指', 'ih': '上证50股指', 'ic': '中证500股指', 'tf': '五债', 't': '十债', 'ts': '二债',
                 'cu': '铜', 'al': '铝', 'zn': '锌', 'pb': '铅', 'ni': '镍', 'sn': '锡', 'au': '黄金', 'ag': '白银',
                 'j': '焦炭', 'jm': '焦煤', 'zc': '动力煤ZC',
                 'rb': '螺纹钢', 'i': '铁矿石', 'hc': '热轧卷板', 'sf': '硅铁', 'sm': '锰硅', 'fg': '玻璃', 'ss': '不锈钢',
                 'wr': '线材', 'ru': '天然橡胶', 'sp': '漂针浆', 'bb': '细木工板', 'fb': '中密度纤维板', 'nr': '20号胶',
                 'fu': '燃料油', 'bu': '石油沥青', 'l': '线型低密度聚乙烯', 'pp': '聚丙烯', 'v': '聚氯乙烯', 'ta': '化纤',
                 'ma': '甲醇MA', 'eg': '乙二醇', 'eb': '苯乙烯', 'ur': '尿素', 'sa': '纯碱', 'pg': '液化石油气',
                 'm': '豆粕', 'y': '豆油', 'oi': '菜籽油', 'a': '黄大豆1号', 'b': '黄大豆2号', 'p': '棕榈油', 'c': '黄玉米',
                 'rm': '菜籽粕', 'cs': '玉米淀粉',
                 'cf': '一号棉花', 'cy': '棉纱', 'sr': '白砂糖', 'wh': '强筋小麦', 'ri': '旱籼稻', 'rr': '粳米',
                 'rs': '油菜籽', 'jr': '粳稻谷', 'lr': '晚籼稻', 'pm': '普通小麦', 'sc': '原油',
                 'ap': '鲜苹果', 'jd': '鲜鸡蛋', 'cj': '干制红枣'}

TRADING_UNIT = {'if': 300, 'ih': 300, 'ic': 200, 'tf': 10000, 't': 10000, 'ts': 20000, 'cu': 5, 'al': 5, 'zn': 5,
                'sc': 1000, 'pb': 5, 'ni': 1, 'sn': 1, 'au': 1000, 'ag': 15, 'j': 100, 'jm': 60, 'zc': 100, 'rb': 10,
                'i': 100, 'hc': 10, 'sf': 5, 'sm': 5, 'wr': 10, 'fu': 10, 'bu': 10, 'ru': 10, 'l': 5, 'pp': 5, 'v': 5,
                'ta': 5, 'ma': 10, 'sp': 10, 'm': 10, 'y': 10, 'oi': 10, 'a': 10, 'b': 10, 'p': 10, 'c': 10, 'rm': 10,
                'cs': 10, 'jd': 10, 'bb': 500, 'fb': 500, 'cf': 5, 'cy': 5, 'sr': 10, 'wh': 20, 'ri': 20, 'jr': 20,
                'lr': 20, 'fg': 20, 'ss': 5, 'nr': 10, 'eg': 10, 'eb': 5, 'ur': 20, 'rr': 10, 'rs': 10, 'ap': 10,
                'cj': 5, 'pm': 50, 'sa': 20, 'pg': 20, }


# ---------------------------------------------------- 基础数据 结束 ----------------------------------------------------


# ---------------------------------------------------- 数据读取 开始 ----------------------------------------------------
def read_statement_files(dir_input=''):
    print('期货交易结算单数据分析\n')
    print('*' * 26 + '使用准备' + '*' * 26 + '\n' +
          '*' + ' ' * 58 + '*' + '\n' +
          '*  请把所有交易结算单txt文件放入当前或指定路径的文件夹中!  *\n' +
          '*' + ' ' * 58 + '*' + '\n' +
          '*' * 60 + '\n')
    if os.path.isdir(dir_input):
        if os.path.exists(dir_input):
            path = dir_input
            folder_name = os.path.basename(path)
        else:
            input('%.19s 错误： %s 路径错误，请检查路径并重试，按任意键退出！' % (datetime.now(), dir_input))
            raise SystemExit()
    else:
        folder = input('输入文件夹名/文件夹路径：') if dir_input == '' else dir_input
        path = os.path.join(BASE_DIR, folder) if not os.path.isdir(folder) else folder
        folder_name = os.path.basename(path)
        if not os.path.exists(path):
            input('%.19s 错误：找不到 %s 文件夹，请检查并重试，按任意键退出！' % (datetime.now(), folder_name))
            raise SystemExit()
    files_list = os.listdir(path)
    statement_data_list = []
    for file in files_list:
        file_dir = os.path.join(path, file)
        if os.path.isfile(file_dir):
            with open(file_dir) as f:
                statement_data_list.append(f.readlines())
    return folder_name, statement_data_list


# ---------------------------------------------------- 数据读取 结束 ----------------------------------------------------


# ---------------------------------------------------- 数据提取 开始 ----------------------------------------------------
def data_extract(source, client_id=''):
    if client_id == '':
        client_id = re.search(r'[^客户号 ClientID：][0-9]+', source[0][8]).group()
    print('%.19s 信息：开始客户号 %s 的结算数据提取' % (datetime.now(), client_id))
    account = pd.DataFrame(columns=['日期', '期初结存', '出入金', '平仓盈亏', '持仓盯市盈亏', '手续费', '期末结存',
                                    '客户权益', '保证金占用', '风险度'])
    date, balance_bf, deposit_withdraw, realized_pl, mtm_pl, commission, balance_cf, client_equity, margin_occupied = [
        list() for _ in range(9)]
    for i in range(len(source)):
        if client_id != re.search(r'[^客户号 ClientID：][0-9]+', source[i][8]).group():
            input('%.19s 错误： %s 结算单客户号不正确，请检查后重试，按任意键退出！' %
                  (datetime.now(), re.search(r'[^日期 Date：][0-9][0-9][0-9][0-9][0-9][0-9][0-9]', source[i][10]).group()))
            raise SystemExit()
        date.append(re.search(r'[^日期 Date：][0-9][0-9][0-9][0-9][0-9][0-9][0-9]', source[i][10]).group())
        balance_bf.append(float(source[i][24][17:45].strip()))
        deposit_withdraw.append(float(source[i][26][25:47].strip()))
        realized_pl.append(float(source[i][28][18:46].strip()))
        mtm_pl.append(float(source[i][30][15:44].strip()))
        commission.append(float(source[i][34][17:46].strip()))
        balance_cf.append(float(source[i][26][65:-1].strip()))
        client_equity.append(float(source[i][30][65:-1].strip()))
        margin_occupied.append(float(source[i][34][70:-1].strip()))
    account['日期'] = pd.to_datetime(date)
    account['期初结存'] = balance_bf
    account['出入金'] = deposit_withdraw
    account['平仓盈亏'] = realized_pl
    account['持仓盯市盈亏'] = mtm_pl
    account['手续费'] = commission
    account['期末结存'] = balance_cf
    account['客户权益'] = client_equity
    account['保证金占用'] = margin_occupied
    account['风险度'] = account['保证金占用'] / account['客户权益']

    sep = re.compile(r'[\|\s|]+')

    transaction_record = pd.DataFrame(columns=['成交日期', '交易所', '品种', '合约', '买/卖', '投/保', '成交价', '手数',
                                               '成交额', '开/平', '手续费', '平仓盈亏', '权利金收支', '成交序号'])
    for i in range(len(source)):
        for j in range(len(source[i])):
            if re.match(r'\s*成交记录 Transaction Record', source[i][j]):
                for l in range(j + 10, len(source[i])):
                    if re.match(r'^-', source[i][l]):
                        break
                    if re.match(r'^\n', source[i][l]):
                        continue
                    row = sep.split(source[i][l])
                    data = {'成交日期': [pd.to_datetime(row[1])], '交易所': [row[2]], '品种': [row[3]], '合约': [row[4]],
                            '买/卖': [row[5]], '投/保': [row[6]], '成交价': [float(row[7])], '手数': [int(row[8])],
                            '成交额': [float(row[9])], '开/平': [row[10]], '手续费': [float(row[11])],
                            '平仓盈亏': [float(row[12])], '权利金收支': [float(row[13])], '成交序号': [int(row[14])]}
                    df = pd.DataFrame(data)
                    transaction_record = transaction_record.append(df)

    position_closed = pd.DataFrame(columns=['平仓日期', '交易所', '品种', '合约', '开仓日期', '买/卖', '手数', '开仓价',
                                            '昨结算', '成交价', '平仓盈亏', '权利金收支', '交易盈亏', '盈亏率'])
    for i in range(len(source)):
        for j in range(len(source[i])):
            if re.match(r'\s*平仓明细 Position Closed', source[i][j]):
                for l in range(j + 10, len(source[i])):
                    if re.match(r'^-', source[i][l]):
                        break
                    if re.match(r'^\n', source[i][l]):
                        continue
                    row = sep.split(source[i][l])
                    price_margin = float(row[10]) - float(row[8]) if row[6] == '卖' else float(row[8]) - float(row[10])
                    data = {'平仓日期': [pd.to_datetime(row[1])], '交易所': [row[2]], '品种': [row[3]], '合约': [row[4]],
                            '开仓日期': [pd.to_datetime(row[5])], '买/卖': [row[6]], '手数': [int(row[7])],
                            '开仓价': [float(row[8])], '昨结算': [float(row[9])], '成交价': [float(row[10])],
                            '平仓盈亏': [float(row[11])], '权利金收支': [float(row[12])],
                            '交易盈亏': [price_margin * int(row[7]) *
                                     TRADING_UNIT[re.sub(r'[^A-Za-z]', '', row[4]).lower()]],
                            '盈亏率': [price_margin / float(row[8])],
                            }
                    df = pd.DataFrame(data)
                    position_closed = position_closed.append(df)

    return client_id, account, transaction_record, position_closed


# ---------------------------------------------------- 数据提取 结束 ----------------------------------------------------


# ---------------------------------------------------- 数据统计 开始 ----------------------------------------------------
def data_statistic(transaction_record, position_closed):
    contracts_analysis = pd.DataFrame(columns=['品种', '合约', '合约平仓盈亏', '合约净盈亏', '交易次数', '交易手数',
                                               '盈利次数', '盈利手数', '交易成功率', '交易盈亏率', '均次盈亏', '均手盈亏',
                                               '成交额'])
    contracts_analysis['合约'] = position_closed.groupby('合约').groups.keys()
    contracts_analysis = contracts_analysis.set_index('合约')
    for index in contracts_analysis.index:
        contracts_analysis.loc[index]['品种'] = CONTRACT_CODE[re.sub(r'[^A-Za-z]', '', index).lower()]
    posc = position_closed.groupby('合约')
    contracts_analysis['合约平仓盈亏'] = posc['交易盈亏'].sum()
    contracts_analysis['合约净盈亏'] = posc['交易盈亏'].sum() - transaction_record.groupby('合约')['手续费'].sum()
    contracts_analysis['交易次数'] = posc.count()
    contracts_analysis['交易手数'] = posc['手数'].sum()
    contracts_analysis['盈利次数'] = posc.apply(lambda x: sum(x['交易盈亏'] > 0))
    contracts_analysis['盈利手数'] = posc.apply(lambda x: sum(x[x['交易盈亏'] > 0]['手数']))
    contracts_analysis['交易成功率'] = round(contracts_analysis['盈利次数'] / contracts_analysis['交易次数'], 4)
    contracts_analysis['交易盈亏率'] = round(contracts_analysis['盈利手数'] / contracts_analysis['交易手数'], 4)
    contracts_analysis['均次盈亏'] = round(contracts_analysis['合约平仓盈亏'] / contracts_analysis['交易次数'], 2)
    contracts_analysis['均手盈亏'] = round(contracts_analysis['合约平仓盈亏'] / contracts_analysis['交易手数'], 2)
    contracts_analysis['成交额'] = transaction_record.groupby('合约')['成交额'].sum()
    contracts_analysis = contracts_analysis.reset_index()

    categories_analysis = pd.DataFrame(columns=['品种', '品种平仓盈亏', '品种净盈亏', '交易次数', '交易手数', '盈利次数',
                                                '盈利手数', '交易成功率', '交易盈亏率', '均次盈亏', '均手盈亏', '成交额'])
    categories_analysis['品种'] = contracts_analysis.groupby('品种').groups.keys()
    categories_analysis = categories_analysis.set_index('品种')
    cata = contracts_analysis.groupby('品种')
    categories_analysis['品种平仓盈亏'] = cata['合约平仓盈亏'].sum()
    categories_analysis['品种净盈亏'] = cata['合约净盈亏'].sum()
    categories_analysis['交易次数'] = cata['交易次数'].sum()
    categories_analysis['交易手数'] = cata['交易手数'].sum()
    categories_analysis['盈利次数'] = cata['盈利次数'].sum()
    categories_analysis['盈利手数'] = cata['盈利手数'].sum()
    categories_analysis['交易成功率'] = round(categories_analysis['盈利次数'] / categories_analysis['交易次数'], 4)
    categories_analysis['交易盈亏率'] = round(categories_analysis['盈利手数'] / categories_analysis['交易手数'], 4)
    categories_analysis['均次盈亏'] = round(categories_analysis['品种平仓盈亏'] / categories_analysis['交易次数'], 2)
    categories_analysis['均手盈亏'] = round(categories_analysis['品种平仓盈亏'] / categories_analysis['交易手数'], 2)
    categories_analysis['成交额'] = cata['成交额'].sum()
    categories_analysis = categories_analysis.reset_index()

    return contracts_analysis, categories_analysis


# ---------------------------------------------------- 数据统计 结束 ----------------------------------------------------


# --------------------------------------------------- 数据格式化 开始 ---------------------------------------------------
def excel_data_format(excel_file):
    wb = load_workbook(excel_file)
    for ws in wb:
        for column_index in [chr(i) for i in range(65, 65 + ws.max_column)]:
            ws.column_dimensions[column_index].auto_size = True
        for data_set in ws.columns:
            header = data_set[0].value
            if '日期' in header:
                for i in range(ws.min_row, ws.max_row):
                    data_set[i].number_format = numbers.FORMAT_DATE_YYYYMMDD2
            elif '风险度' in header:
                for i in range(ws.min_row, ws.max_row):
                    data_set[i].number_format = numbers.FORMAT_PERCENTAGE_00
            elif '盈亏率' in header or '成功率' in header:
                for i in range(ws.min_row, ws.max_row):
                    data_set[i].number_format = numbers.FORMAT_PERCENTAGE_00
            elif '盈亏' in header or '结存' in header or '权益' in header or '保证金' in header or '出入金' in header or '成交额' in header:
                for i in range(ws.min_row, ws.max_row):
                    data_set[i].number_format = numbers.BUILTIN_FORMATS[39]
    wb.save(excel_file)
    wb.close()


# --------------------------------------------------- 数据格式化 结束 ---------------------------------------------------


# ---------------------------------------------------- 生成图表 开始 ----------------------------------------------------
def excel_create_chart(excel_file):
    """
    生成图表
    :param excel_file:
    :return:
    """
    wb = load_workbook(excel_file)
    # 进行权益走势图渲染
    account_sheet = wb['账户统计']
    account_chart_sheet = wb.create_chartsheet(title='权益走势图')
    dates = Reference(account_sheet, min_col=1, min_row=2, max_row=len(account_sheet['A']))
    chart1 = LineChart(varyColors=True)
    data1 = Reference(account_sheet, min_col=8, min_row=1, max_row=len(account_sheet['H']))
    chart1.add_data(data1, titles_from_data=True)
    chart1.y_axis = NumericAxis(axId=100, title='权益', majorTickMark='out')
    chart1.legend = None
    chart1.set_categories(dates)
    chart2 = BarChart()
    data2 = Reference(account_sheet, min_col=10, min_row=1, max_row=len(account_sheet['J']))
    chart2.add_data(data2, titles_from_data=True)
    chart2.y_axis = NumericAxis(axId=200, title='风险度', majorGridlines=None, majorTickMark='out', crosses='max')
    chart2.legend = None
    chart2.set_categories(dates)
    chart1 += chart2
    chart1.x_axis.majorTickMark = 'out'
    chart1.x_axis.tickLblSkip = 10
    chart1.x_axis.tickMarkSkip = 10
    chart1.x_axis.noMultiLvlLbl = True
    chart1.x_axis.number_format = 'yyyy-mm-dd'
    account_chart_sheet.add_chart(chart1)
    # 进行交易分布图的渲染
    categories_analysis_sheet = wb['交易分析(按品种)']
    trading_frequency_analysis_sheet = wb.create_chartsheet(title='交易分布图')
    labels = Reference(categories_analysis_sheet, min_col=1, min_row=2, max_row=len(categories_analysis_sheet['A']))
    chart3 = PieChart(varyColors=True)
    chart3.style = 34
    data3 = Reference(categories_analysis_sheet, min_col=4, min_row=2, max_row=len(categories_analysis_sheet['D']))
    chart3.add_data(data3)
    chart3.set_categories(labels)
    chart3.legend = None
    # chart3.series[0].data_points = [DataPoint(idx=i, explosion=8) for i in range(len(categories_analysis_sheet['D']) - 1)]
    chart3.series[0].dLbls = DataLabelList(dLblPos='bestFit', showPercent=True, showCatName=True, showVal=True,
                                           showLeaderLines=True)
    chart3.layout = Layout(manualLayout=ManualLayout(x=0, y=0, h=0.75, w=0.75, xMode='factor', yMode='factor'))
    trading_frequency_analysis_sheet.add_chart(chart3)
    # 进行品种盈亏图的渲染
    categories_win_and_loss_chart_sheet = wb.create_chartsheet(title='品种盈亏图')
    chart4 = BarChart(barDir='col')
    chart4.style = 18
    data4 = Reference(categories_analysis_sheet, min_col=2, min_row=2, max_row=len(categories_analysis_sheet['B']))
    chart4.add_data(data4)
    chart4.set_categories(labels)
    chart4.legend = None
    chart4.series[0].dLbls = DataLabelList(showVal=True)
    chart4.y_axis = NumericAxis(title='平仓盈亏', majorTickMark='out', minorTickMark='out')
    categories_win_and_loss_chart_sheet.add_chart(chart4)
    # 进行交易盈亏图的渲染
    trading_win_and_loss_chart_sheet = wb.create_chartsheet(title='交易盈亏图')
    chart5 = RadarChart()
    chart5.style = 24
    data5 = Reference(categories_analysis_sheet, min_col=8, max_col=9, min_row=1, max_row=categories_analysis_sheet.max_row)
    chart5.add_data(data5, titles_from_data=True)
    chart5.set_categories(labels)
    trading_win_and_loss_chart_sheet.add_chart(chart5)
    # 图表保存
    wb.save(excel_file)
    wb.close()


# ---------------------------------------------------- 生成图表 结束 ----------------------------------------------------


# -------------------------------------------------- 生成excel文件 开始 -------------------------------------------------
def output_excel(account, transaction_record, position_closed, contracts_analysis, categories_analysis, client_id=''):
    try:
        excel_file_name = '%s%s%s交易统计.xlsx' % (BASE_DIR, '\\' if sys.platform == 'win32' else '/', client_id)
        with pd.ExcelWriter(excel_file_name) as writer:
            account.to_excel(writer, sheet_name='账户统计', encoding='ansi', index=None)
            transaction_record.to_excel(writer, sheet_name='交易记录', encoding='ansi', index=None)
            position_closed.to_excel(writer, sheet_name='平仓明细', encoding='ansi', index=None)
            contracts_analysis.to_excel(writer, sheet_name='交易分析(按合约)', encoding='ansi', index=None)
            categories_analysis.to_excel(writer, sheet_name='交易分析(按品种)', encoding='ansi', index=None)
        excel_data_format(excel_file_name)
        print('%.19s 信息：Excel分析结果已生成' % datetime.now())
        excel_create_chart(excel_file_name)
        input('%.19s 信息：Excel分析图表已生成' % datetime.now())
    except PermissionError:
        input('%.19s 错误：分析结果写入被拒绝，请检查文件是否已打开，按任意键退出！' % datetime.now())
        raise SystemExit()


# -------------------------------------------------- 生成excel文件 结束 -------------------------------------------------


# ---------------------------------------------------- 终端命令 开始 ----------------------------------------------------
def main(argv):
    client_id = ''
    files_folder = ''
    try:
        opts, args = getopt.getopt(argv, "hd:i:", ["dir=", "id="])
    except getopt.GetoptError:
        print('参数选项:\n-d/--dir <settlement statement files\' folder  结算单文件夹路径>\n-i/--id <client id  客户号>')
        sys.exit(2)
    if len(opts) != 0:
        for opt, arg in opts:
            if opt == '-h':
                print('参数选项:\n-d/--dir <settlement statement files\' folder  结算单文件夹路径>\n-i/--id <client id  客户号>')
                sys.exit()
            elif opt in ("-d", "--dir"):
                files_folder = arg
            elif opt in ("-i", "--id"):
                client_id = arg
    files_folder_name, statement_list = read_statement_files(files_folder)
    print('%.19s 信息：已读取 %s 文件夹' % (datetime.now(), files_folder_name))
    client_id, account, transaction_record, position_closed = data_extract(statement_list, client_id=client_id)
    print('%.19s 信息：所有结算单数据已提取' % datetime.now())
    contracts_analysis, categories_analysis = data_statistic(transaction_record, position_closed)
    output_excel(account, transaction_record, position_closed, contracts_analysis, categories_analysis,
                 client_id=client_id)


if __name__ == '__main__':
    main(sys.argv[1:])
# ---------------------------------------------------- 终端命令 结束 ----------------------------------------------------
